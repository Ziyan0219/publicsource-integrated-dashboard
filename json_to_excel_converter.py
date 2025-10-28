#!/usr/bin/env python3
"""
Convert stories.json to Excel format for easy review
"""
import json
import pandas as pd
from pathlib import Path
from datetime import datetime

def convert_json_to_excel():
    """Convert stories.json to Excel with multiple sheets"""

    # Load JSON data
    stories_path = Path("frontend/src/data/stories.json")
    with open(stories_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    stories = data.get('stories', [])
    filters = data.get('filters', {})

    print(f"Loading {len(stories)} stories from JSON...")

    # Create main stories DataFrame
    stories_df = pd.DataFrame(stories)

    # Reorder columns for better readability
    column_order = [
        'id', 'title', 'author', 'date', 'url',
        'umbrella', 'geographic_area', 'neighborhoods',
        'social_abstract'
    ]

    # Only include columns that exist
    available_columns = [col for col in column_order if col in stories_df.columns]
    stories_df = stories_df[available_columns]

    # Create Excel writer
    output_file = f"stories_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print(f"Creating Excel file: {output_file}")

    # Calculate date range safely
    date_range = 'N/A'
    if 'date' in stories_df:
        try:
            date_series = pd.to_datetime(stories_df['date'], errors='coerce')
            date_series = date_series.dropna()
            if not date_series.empty:
                date_range = f"{date_series.min().strftime('%Y-%m-%d')} to {date_series.max().strftime('%Y-%m-%d')}"
        except:
            pass

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: All Stories
        stories_df.to_excel(writer, sheet_name='All Stories', index=False)

        # Sheet 2: Summary Statistics
        summary_data = {
            'Metric': [
                'Total Stories',
                'Unique Umbrellas',
                'Unique Geographic Areas',
                'Unique Neighborhoods',
                'Date Range',
                'Stories with Social Abstract',
                'Average Abstract Length'
            ],
            'Value': [
                len(stories),
                len(filters.get('umbrellas', [])),
                len(filters.get('geographic_areas', [])),
                len(filters.get('neighborhoods', [])),
                date_range,
                stories_df['social_abstract'].notna().sum() if 'social_abstract' in stories_df else 0,
                int(stories_df['social_abstract'].str.len().mean()) if 'social_abstract' in stories_df else 0
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Sheet 3: Umbrellas
        if filters.get('umbrellas'):
            umbrellas_df = pd.DataFrame({
                'Umbrella Category': sorted(filters['umbrellas']),
                'Story Count': [
                    stories_df['umbrella'].str.contains(umb, na=False, case=False, regex=False).sum()
                    for umb in sorted(filters['umbrellas'])
                ]
            })
            umbrellas_df.to_excel(writer, sheet_name='Umbrellas', index=False)

        # Sheet 4: Geographic Areas
        if filters.get('geographic_areas'):
            geo_areas_df = pd.DataFrame({
                'Geographic Area': sorted(filters['geographic_areas']),
                'Story Count': [
                    stories_df['geographic_area'].str.contains(area, na=False, case=False, regex=False).sum()
                    for area in sorted(filters['geographic_areas'])
                ]
            })
            geo_areas_df.to_excel(writer, sheet_name='Geographic Areas', index=False)

        # Sheet 5: Neighborhoods (top 50 by frequency)
        if filters.get('neighborhoods'):
            all_neighborhoods = filters['neighborhoods']
            neighborhood_counts = []
            for neigh in all_neighborhoods:
                count = stories_df['neighborhoods'].str.contains(neigh, na=False, case=False, regex=False).sum()
                neighborhood_counts.append({'Neighborhood': neigh, 'Story Count': count})

            neighborhoods_df = pd.DataFrame(neighborhood_counts)
            neighborhoods_df = neighborhoods_df.sort_values('Story Count', ascending=False).head(100)
            neighborhoods_df.to_excel(writer, sheet_name='Top 100 Neighborhoods', index=False)

        # Sheet 6: Stories by Year
        if 'date' in stories_df:
            try:
                stories_df['year'] = pd.to_datetime(stories_df['date'], errors='coerce', utc=True).dt.year
                yearly_stats = stories_df[stories_df['year'].notna()].groupby('year').agg({
                    'id': 'count',
                    'umbrella': lambda x: x.nunique(),
                    'neighborhoods': lambda x: x.notna().sum()
                }).reset_index()
                yearly_stats.columns = ['Year', 'Total Stories', 'Unique Umbrellas', 'Stories with Neighborhoods']
                yearly_stats.to_excel(writer, sheet_name='Yearly Breakdown', index=False)
            except Exception as e:
                print(f"  - Warning: Could not create yearly breakdown: {e}")

        # Sheet 7: Stories without Geographic Data
        missing_geo = stories_df[
            (stories_df['umbrella'].isna() | (stories_df['umbrella'] == '')) |
            (stories_df['geographic_area'].isna() | (stories_df['geographic_area'] == '')) |
            (stories_df['neighborhoods'].isna() | (stories_df['neighborhoods'] == ''))
        ].copy()

        if not missing_geo.empty:
            missing_geo.to_excel(writer, sheet_name='Missing Geographic Data', index=False)
            print(f"  - Found {len(missing_geo)} stories with missing geographic data")

    # Auto-adjust column widths (approximate)
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 80)  # Cap at 80 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

    wb.save(output_file)

    print(f"\n[SUCCESS] Excel file created successfully!")
    print(f"File location: {Path(output_file).absolute()}")
    print(f"\nSheets included:")
    print(f"  1. All Stories - Complete dataset with all fields")
    print(f"  2. Summary - High-level statistics")
    print(f"  3. Umbrellas - All umbrella categories with story counts")
    print(f"  4. Geographic Areas - All geographic areas with story counts")
    print(f"  5. Top 100 Neighborhoods - Most mentioned neighborhoods")
    print(f"  6. Yearly Breakdown - Statistics by year")
    print(f"  7. Missing Geographic Data - Stories needing review")

    return output_file

if __name__ == "__main__":
    convert_json_to_excel()
